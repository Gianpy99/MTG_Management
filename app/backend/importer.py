"""Workbook / CSV import logic.

The Excel workbook is the authoritative seed. Import is transactional and
reports validation issues rather than silently changing data.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from sqlalchemy.orm import Session

from models import Card, ImportLog

# Canonical header -> Card attribute mapping (order-independent, case-insensitive).
HEADER_MAP = {
    "set": "set_name",
    "card name": "card_name",
    "collector number": "collector_number",
    "rarity": "rarity",
    "colour": "colour",
    "color": "colour",
    "mana cost": "mana_cost",
    "card type": "card_type",
    "subtype": "subtype",
    "power": "power",
    "toughness": "toughness",
    "oracle text / ability": "oracle_text",
    "oracle text": "oracle_text",
    "legendary?": "legendary",
    "creature type": "creature_type",
    "ring tempts you?": "ring_tempts",
    "food?": "food",
    "treasure?": "treasure",
    "ring / the one ring synergy?": "ring_synergy",
    "aragorn synergy (1-5)": "aragorn_synergy",
    "gandalf synergy (1-5)": "gandalf_synergy",
    "fellowship / legends synergy (1-5)": "fellowship_synergy",
    "commander role": "commander_role",
    "owned?": "_owned",
    "notes": "notes",
}

BOOL_FIELDS = {"legendary", "ring_tempts", "food", "treasure", "ring_synergy"}
INT_FIELDS = {"aragorn_synergy", "gandalf_synergy", "fellowship_synergy"}


def _to_bool(value: str) -> bool:
    return str(value).strip().lower() in {"yes", "y", "true", "1"}


def _to_int(value: str) -> int:
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return 0


@dataclass
class ImportResult:
    added: int = 0
    updated: int = 0
    unchanged: int = 0
    rejected: int = 0
    issues: list[str] = field(default_factory=list)

    def summary(self) -> str:
        return (
            f"added={self.added} updated={self.updated} "
            f"unchanged={self.unchanged} rejected={self.rejected} "
            f"issues={len(self.issues)}"
        )


def _rows_from_csv(text: str) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Collection"] if "Collection" in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    out: list[dict[str, str]] = []
    for r in rows:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({headers[i]: ("" if c is None else str(c)) for i, c in enumerate(r) if i < len(headers)})
    return out


def import_rows(db: Session, rows: list[dict[str, str]]) -> ImportResult:
    result = ImportResult()
    # Track identities seen in THIS batch so duplicate rows are merged into the
    # same Card instead of violating the unique constraint (reported as issues).
    pending: dict[tuple[str, str, str], Card] = {}
    for idx, raw in enumerate(rows, start=2):
        # Normalise keys to the mapped attributes.
        mapped: dict[str, str] = {}
        for key, value in raw.items():
            attr = HEADER_MAP.get(str(key).strip().lower())
            if attr:
                mapped[attr] = value if value is not None else ""

        name = (mapped.get("card_name") or "").strip()
        set_name = (mapped.get("set_name") or "").strip()
        if not name:
            result.rejected += 1
            result.issues.append(f"row {idx}: missing card name")
            continue
        if not set_name:
            result.rejected += 1
            result.issues.append(f"row {idx}: '{name}' missing set")
            continue

        owned = _to_bool(mapped.pop("_owned", "No")) if "_owned" in mapped else False
        collector = (mapped.get("collector_number") or "").strip()

        key = (set_name, collector, name)
        existing = pending.get(key)
        if existing is None:
            existing = (
                db.query(Card)
                .filter(Card.set_name == set_name, Card.card_name == name, Card.collector_number == collector)
                .one_or_none()
            )
        else:
            result.issues.append(f"row {idx}: duplicate identity '{name}' ({set_name} {collector}) merged")

        target = existing or Card(set_name=set_name, card_name=name, collector_number=collector)
        changed = existing is None
        for attr, value in mapped.items():
            if attr in BOOL_FIELDS:
                new_val = _to_bool(value)
            elif attr in INT_FIELDS:
                new_val = _to_int(value)
            else:
                new_val = (value or "").strip()
            if getattr(target, attr, None) != new_val:
                setattr(target, attr, new_val)
                changed = True

        if existing is None:
            target.quantity = 1 if owned else 0
            db.add(target)
            pending[key] = target
            result.added += 1
        elif changed:
            result.updated += 1
        else:
            result.unchanged += 1

    db.add(ImportLog(summary=result.summary()))
    db.commit()
    return result


def import_file(db: Session, filename: str, data: bytes) -> ImportResult:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(data)
    else:
        rows = _rows_from_csv(data.decode("utf-8-sig"))
    return import_rows(db, rows)
